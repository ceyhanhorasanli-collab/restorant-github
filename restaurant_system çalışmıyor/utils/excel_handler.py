"""
Excel Handler - Excel dosya işlemleri için güvenli fonksiyonlar
"""

import os
import pandas as pd
from config import OPENPYXL_AVAILABLE

if OPENPYXL_AVAILABLE:
    from openpyxl import Workbook, load_workbook


def safe_excel_write(file_path, data, headers, sheet_name="Sheet1"):
    """
    Güvenli Excel yazma fonksiyonu - doğrudan openpyxl kullanır
    Pandas DataFrame problemlerini bypass eder
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl kütüphanesi gerekli!")
    
    try:
        # Yeni workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Başlıkları yaz
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Verileri yaz
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Dosyayı kaydet
        wb.save(file_path)
        return True
        
    except Exception as e:
        print(f"❌ Excel yazma hatası: {e}")
        return False


def safe_excel_append(file_path, data, headers):
    """
    Güvenli Excel'e ekleme fonksiyonu - mevcut dosyayı okur, veri ekler, yeniden yazar
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl kütüphanesi gerekli!")
    
    try:
        # Dosya var mı kontrol et
        if os.path.exists(file_path):
            # Mevcut dosyayı oku
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Mevcut verileri topla
            existing_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Boş satırları atla
                    existing_data.append(list(row))
            
            # Yeni veriyi ekle
            combined_data = existing_data + data
        else:
            # Dosya yoksa sadece yeni veriyi kullan
            combined_data = data
        
        # Tüm veriyi tekrar yaz
        return safe_excel_write(file_path, combined_data, headers)
        
    except Exception as e:
        print(f"❌ Excel ekleme hatası: {e}")
        return False


def safe_excel_replace_meal(file_path, meal_name, new_data, headers):
    """
    Belirli bir yemeğin reçetesini değiştir (aynı isimli yemeği sil, yenisini ekle)
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl kütüphanesi gerekli!")
    
    try:
        # Dosya var mı kontrol et
        if os.path.exists(file_path):
            # Mevcut dosyayı oku
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Mevcut verileri topla (aynı yemek adı olanları hariç)
            existing_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 0 and row[0] != meal_name:  # İlk sütun yemek adı
                    existing_data.append(list(row))
        else:
            existing_data = []
        
        # Yeni veriyi ekle
        combined_data = existing_data + new_data
        
        # Tüm veriyi tekrar yaz
        return safe_excel_write(file_path, combined_data, headers)
        
    except Exception as e:
        print(f"❌ Excel replace hatası: {e}")
        return False


def create_sample_excels(base_path):
    """Örnek Excel dosyalarını oluştur.

    DİKKAT: Mevcut `.xlsx` dosyalarının üzerine yazılmaz. Eğer hedef dosya zaten
    mevcutsa aynı isimle `.sample.xlsx` veya `.sampleN.xlsx` şeklinde yeni bir örnek
    dosya oluşturulur. Bu sayede kullanıcı verileri korunur (geçici mantık).
    """
    try:
        def _unique_sample_path(path):
            base, ext = os.path.splitext(path)
            sample_path = f"{base}.sample{ext}"
            if not os.path.exists(sample_path):
                return sample_path
            # Eğer .sample.xlsx zaten varsa, numaralandır
            i = 1
            while True:
                candidate = f"{base}.sample{i}{ext}"
                if not os.path.exists(candidate):
                    return candidate
                i += 1

        yemek_cols = ["yemek adı", "porsiyon adt", "ürün", "miktar"]
        yemek_data = [
            ["Kıymalı Makarna", 4, "Kıyma", 0.5],
            ["Kıymalı Makarna", 4, "Makarna", 0.4],
            ["Mercimek Çorbası", 6, "Mercimek", 0.3],
            ["Mercimek Çorbası", 6, "Soğan", 0.05],
        ]
        df_y = pd.DataFrame(yemek_data, columns=yemek_cols)
        yemek_path = os.path.join(base_path, "yemekler.xlsx")
        if os.path.exists(yemek_path):
            target_y = _unique_sample_path(yemek_path)
            df_y.to_excel(target_y, index=False)
            print(f"ℹ️ Mevcut 'yemekler.xlsx' korundu; örnek oluşturuldu: {os.path.basename(target_y)}")
        else:
            df_y.to_excel(yemek_path, index=False)

        urun_cols = ["Ürün Adı", "Tarih", "Miktar", "Alış Fiyatı (TL)"]
        urun_data = [
            ["Kıyma", "2025-06-01", 2.0, 520.0],
            ["Kıyma", "2025-07-10", 1.5, 330.0],
            ["Kıyma", "2025-08-05", 1.0, 150.0],
            ["Makarna", "2025-08-02", 5.0, 250.0],
            ["Mercimek", "2025-08-01", 4.0, 200.0],
            ["Soğan", "2025-08-03", 10.0, 120.0],
        ]
        df_u = pd.DataFrame(urun_data, columns=urun_cols)
        urun_path = os.path.join(base_path, "urunler.xlsx")
        if os.path.exists(urun_path):
            target_u = _unique_sample_path(urun_path)
            df_u.to_excel(target_u, index=False)
            print(f"ℹ️ Mevcut 'urunler.xlsx' korundu; örnek oluşturuldu: {os.path.basename(target_u)}")
        else:
            df_u.to_excel(urun_path, index=False)

        print("✅ Örnek excel dosyaları oluşturuldu (mevcutlar korunarak).")
    except Exception:
        print("❌ Örnek excel oluşturulurken hata:")
        import traceback
        traceback.print_exc()
